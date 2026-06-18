"""
generate_report.py - Generates a professional Appium Mobile E2E Test Report in .xlsx format
"""
import os
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference

def generate_excel_report(results, output_dir=None):
    """Generate the .xlsx Appium test report from collected results."""

    if output_dir is None:
        output_dir = os.path.dirname(os.path.abspath(__file__))

    timestamp = datetime.now().strftime("%Y-%m-%dT%H-%M-%S")
    filename = f"Appium_Test_Report_HistoQuanta_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)

    wb = Workbook()

    # ── Styles ──
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="8F3E97", end_color="8F3E97", fill_type="solid") # Purple color for Mobile/Appium
    title_font = Font(name="Calibri", bold=True, size=18, color="8F3E97")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    pass_font = Font(name="Calibri", bold=True, color="006100")
    fail_font = Font(name="Calibri", bold=True, color="9C0006")
    normal_font = Font(name="Calibri", size=11)
    border_thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    alt_fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_align = Alignment(wrap_text=True, vertical="top")

    # ── Compute Stats ──
    total = len(results)
    passed = sum(1 for r in results if r["status"] == "PASS")
    failed = sum(1 for r in results if r["status"] == "FAIL")
    pass_rate = round((passed / total * 100), 1) if total > 0 else 0
    total_exec_time = sum(r.get("execution_time", 0) for r in results)

    # Modules breakdown
    modules = {}
    for r in results:
        mod = r.get("module", "Unknown")
        if mod not in modules:
            modules[mod] = {"total": 0, "passed": 0, "failed": 0}
        modules[mod]["total"] += 1
        if r["status"] == "PASS":
            modules[mod]["passed"] += 1
        else:
            modules[mod]["failed"] += 1

    # ═══════════════════════════════════════════════
    #  SHEET 1: SUMMARY
    # ═══════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.sheet_properties.tabColor = "8F3E97"

    # Title
    ws1.merge_cells("A1:H1")
    ws1["A1"] = "HistoQuanta iOS - Appium Mobile E2E Test Report"
    ws1["A1"].font = title_font
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 45

    ws1.merge_cells("A2:H2")
    ws1["A2"] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    ws1["A2"].font = Font(name="Calibri", size=11, color="666666")
    ws1["A2"].alignment = Alignment(horizontal="center")
    ws1.row_dimensions[2].height = 25

    # Summary Table
    summary_row = 4
    summary_data = [
        ("Metric", "Value"),
        ("Application", "HistoQuanta Mobile iOS App (HistoQuanta 4)"),
        ("Test Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Target Platform", "iOS (Appium Mobile Automation)"),
        ("Device/Simulator", "iPhone 15 (XCUITest)"),
        ("Total Test Cases", str(total)),
        ("Passed", str(passed)),
        ("Failed", str(failed)),
        ("Pass Rate", f"{pass_rate}%"),
        ("Total Execution Time", f"{round(total_exec_time, 1)}s"),
        ("Test Framework", "Pytest + Appium Python Client"),
        ("Report Generator", "openpyxl"),
    ]

    for i, (label, value) in enumerate(summary_data):
        row = summary_row + i
        ws1.cell(row=row, column=1, value=label)
        ws1.cell(row=row, column=2, value=value)
        if i == 0:
            ws1.cell(row=row, column=1).font = header_font
            ws1.cell(row=row, column=1).fill = header_fill
            ws1.cell(row=row, column=2).font = header_font
            ws1.cell(row=row, column=2).fill = header_fill
        else:
            ws1.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=11)
            ws1.cell(row=row, column=2).font = normal_font
            if label == "Passed":
                ws1.cell(row=row, column=2).fill = pass_fill
                ws1.cell(row=row, column=2).font = pass_font
            elif label == "Failed":
                ws1.cell(row=row, column=2).fill = fail_fill
                ws1.cell(row=row, column=2).font = fail_font
        ws1.cell(row=row, column=1).border = border_thin
        ws1.cell(row=row, column=2).border = border_thin

    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 45

    # Pie Chart for Pass/Fail
    pie_row = summary_row + len(summary_data) + 2
    ws1.cell(row=pie_row, column=1, value="Status").font = Font(bold=True, size=11)
    ws1.cell(row=pie_row, column=2, value="Count").font = Font(bold=True, size=11)
    ws1.cell(row=pie_row + 1, column=1, value="PASS")
    ws1.cell(row=pie_row + 1, column=2, value=passed)
    ws1.cell(row=pie_row + 2, column=1, value="FAIL")
    ws1.cell(row=pie_row + 2, column=2, value=failed)

    chart = PieChart()
    chart.title = "Test Results Distribution"
    chart.style = 10
    chart.width = 14
    chart.height = 10
    data_ref = Reference(ws1, min_col=2, min_row=pie_row, max_row=pie_row + 2)
    cats_ref = Reference(ws1, min_col=1, min_row=pie_row + 1, max_row=pie_row + 2)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    from openpyxl.chart.series import DataPoint
    # Purple theme for PASS/FAIL charts
    pt0 = DataPoint(idx=0)
    pt0.graphicalProperties.solidFill = "22C55E"
    chart.series[0].data_points.append(pt0)
    pt1 = DataPoint(idx=1)
    pt1.graphicalProperties.solidFill = "EF4444"
    chart.series[0].data_points.append(pt1)
    ws1.add_chart(chart, f"D{summary_row}")

    # ═══════════════════════════════════════════════
    #  SHEET 2: DETAILED RESULTS
    # ═══════════════════════════════════════════════
    ws2 = wb.create_sheet("Detailed Results")
    ws2.sheet_properties.tabColor = "A855F7"

    headers = [
        "S.No", "Test ID", "Module", "Test Case", "Description",
        "Test Steps", "Expected Result", "Actual Result",
        "Status", "Exec Time (s)", "Timestamp"
    ]

    col_widths = [6, 10, 18, 30, 40, 45, 35, 40, 10, 12, 20]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border_thin
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    ws2.row_dimensions[1].height = 30
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for row_idx, result in enumerate(results, 2):
        row_data = [
            row_idx - 1,
            result.get("test_id", ""),
            result.get("module", ""),
            result.get("test_case", ""),
            result.get("description", ""),
            result.get("steps", ""),
            result.get("expected_result", ""),
            result.get("actual_result", ""),
            result.get("status", ""),
            result.get("execution_time", 0),
            result.get("timestamp", ""),
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.font = normal_font
            cell.border = border_thin
            cell.alignment = wrap_align

            if row_idx % 2 == 0:
                cell.fill = alt_fill

        status_cell = ws2.cell(row=row_idx, column=9)
        status_cell.alignment = center_align
        if result.get("status") == "PASS":
            status_cell.fill = pass_fill
            status_cell.font = pass_font
        elif result.get("status") == "FAIL":
            status_cell.fill = fail_fill
            status_cell.font = fail_font

        ws2.cell(row=row_idx, column=1).alignment = center_align
        ws2.cell(row=row_idx, column=10).alignment = center_align
        ws2.row_dimensions[row_idx].height = 55

    # ═══════════════════════════════════════════════
    #  SHEET 3: MODULE SUMMARY
    # ═══════════════════════════════════════════════
    ws3 = wb.create_sheet("Module Summary")
    ws3.sheet_properties.tabColor = "16A34A"

    mod_headers = ["S.No", "Module", "Total Tests", "Passed", "Failed", "Pass Rate (%)"]
    mod_widths = [6, 25, 14, 12, 12, 14]

    for col_idx, (header, width) in enumerate(zip(mod_headers, mod_widths), 1):
        cell = ws3.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border_thin
        ws3.column_dimensions[get_column_letter(col_idx)].width = width

    ws3.row_dimensions[1].height = 30

    for i, (mod_name, stats) in enumerate(modules.items(), 2):
        rate = round(stats["passed"] / stats["total"] * 100, 1) if stats["total"] > 0 else 0
        row_data = [i - 1, mod_name, stats["total"], stats["passed"], stats["failed"], rate]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws3.cell(row=i, column=col_idx, value=value)
            cell.font = normal_font
            cell.border = border_thin
            cell.alignment = center_align
            if i % 2 == 0:
                cell.fill = alt_fill

        pass_cell = ws3.cell(row=i, column=4)
        fail_cell = ws3.cell(row=i, column=5)
        if stats["passed"] > 0:
            pass_cell.fill = pass_fill
            pass_cell.font = pass_font
        if stats["failed"] > 0:
            fail_cell.fill = fail_fill
            fail_cell.font = fail_font

        ws3.row_dimensions[i].height = 25

    # Total row
    total_row = len(modules) + 2
    ws3.cell(row=total_row, column=1, value="").border = border_thin
    ws3.cell(row=total_row, column=2, value="TOTAL").font = Font(name="Calibri", bold=True, size=12)
    ws3.cell(row=total_row, column=2).border = border_thin
    ws3.cell(row=total_row, column=3, value=total).font = Font(bold=True, size=12)
    ws3.cell(row=total_row, column=3).border = border_thin
    ws3.cell(row=total_row, column=3).alignment = center_align
    ws3.cell(row=total_row, column=4, value=passed).font = pass_font
    ws3.cell(row=total_row, column=4).fill = pass_fill
    ws3.cell(row=total_row, column=4).border = border_thin
    ws3.cell(row=total_row, column=4).alignment = center_align
    ws3.cell(row=total_row, column=5, value=failed).font = fail_font
    ws3.cell(row=total_row, column=5).fill = fail_fill
    ws3.cell(row=total_row, column=5).border = border_thin
    ws3.cell(row=total_row, column=5).alignment = center_align
    ws3.cell(row=total_row, column=6, value=f"{pass_rate}%").font = Font(bold=True, size=12)
    ws3.cell(row=total_row, column=6).border = border_thin
    ws3.cell(row=total_row, column=6).alignment = center_align

    wb.save(filepath)
    print(f"\n{'='*60}")
    print(f"  APPIUM TEST REPORT GENERATED SUCCESSFULLY!")
    print(f"{'='*60}")
    print(f"  File: {filepath}")
    print(f"  Total: {total} | Passed: {passed} | Failed: {failed}")
    print(f"  Pass Rate: {pass_rate}%")
    print(f"  Execution Time: {round(total_exec_time, 1)}s")
    print(f"{'='*60}\n")

    return filepath
